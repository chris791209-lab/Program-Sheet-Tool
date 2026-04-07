import streamlit as st
import pandas as pd
import io
import os
import xlsxwriter
import zipfile
import tempfile
import re 
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from PIL import Image

# --- 網頁介面設定 ---
st.set_page_config(page_title="Program Sheet 生成器", layout="centered")
st.title("🚀 Program Sheet 自動生成器")
st.markdown("請依序完成以下步驟，系統將瞬間為您篩選資料並排版整合！")

# ========================================================
# 📄 步驟 1：上傳商品資料
# ========================================================
st.markdown("### 📄 步驟 1：上傳商品資料")
uploaded_data = st.file_uploader("請上傳包含完整報價、工廠等資訊的 Excel 檔 (.xlsm / .xlsx)", type=["xlsm", "xlsx"])

if uploaded_data is not None:
    st.success("✅ 資料檔已讀取！")
    
    try:
        uploaded_data.seek(0)
        df = pd.read_excel(uploaded_data, sheet_name="Data", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        uploaded_data.seek(0)
        df = pd.read_excel(uploaded_data, engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        
    if 'DPCI' in df.columns:
        df = df.dropna(subset=['DPCI'])

    # ========================================================
    # 🗂️ 步驟 2：資料動態篩選
    # ========================================================
    st.markdown("### 🗂️ 步驟 2：選擇產出類別 (部門分工用)")
    st.info("💡 系統已鎖定分類欄位。若只需產出特定類別供部門分工，請在此篩選；若留空則會產出全部商品。")
    
    target_cols = ['Subclass Name', 'Factory Name']
    available_cols = [c for c in target_cols if c in df.columns]
    
    if available_cols:
        filter_col = st.selectbox("📌 1. 請確認用來分類的「欄位名稱」：", available_cols)
    else:
        st.warning("⚠️ 檔案中找不到 'Subclass Name' 或 'Factory Name'，已顯示所有欄位供您選擇。")
        filter_col = st.selectbox("📌 1. 請確認用來分類的「欄位名稱」：", df.columns)

    unique_vals = sorted([str(x).strip() for x in df[filter_col].dropna().unique() if str(x).strip() != ''])
    selected_categories = st.multiselect(f"🎯 2. 請選擇要輸出的【{filter_col}】(可多選)：", unique_vals)

    if selected_categories:
        df_filtered = df[df[filter_col].astype(str).str.strip().isin(selected_categories)].copy()
        st.write(f"📊 目前已篩選：將產出 **{len(df_filtered)}** 筆商品 (原始總數 {len(df)} 筆)")
    else:
        df_filtered = df.copy()
        st.write(f"📊 目前未套用篩選：將產出 **全部 {len(df_filtered)}** 筆商品")

    # ========================================================
    # 🖼️ 步驟 3：上傳圖片來源
    # ========================================================
    st.markdown("### 🖼️ 步驟 3：上傳圖片來源 (二擇一)")
    img_option = st.radio("請選擇您的圖片提供方式：", ["📁 上傳 ZIP 壓縮檔 (內含 JPG/PNG)", "📊 上傳有縮圖的 Excel 檔 (自動萃取)"])

    uploaded_img = None
    if img_option == "📁 上傳 ZIP 壓縮檔 (內含 JPG/PNG)":
        uploaded_img = st.file_uploader("請上傳 .zip 圖片壓縮檔", type=["zip"])
    else:
        uploaded_img = st.file_uploader("請上傳含有 Thumbnail (縮圖) 欄位的 Excel 檔", type=["xlsx", "xlsm"])

    # ========================================================
    # ⚙️ 步驟 4：執行生成
    # ========================================================
    st.markdown("### ⚙️ 步驟 4：執行生成")
    if st.button("✨ 生成 Program Sheet"):
        
        if df_filtered.empty:
            st.error("⚠️ 篩選後的資料為空，無法生成！請重新調整您的篩選條件。")
            st.stop()
            
        with st.spinner("正在為您進行排版與處理圖片，請稍候 (若包含萃取圖片可能需要數十秒)..."):
            temp_dir = tempfile.mkdtemp()
            
            if uploaded_img is not None:
                if img_option == "📁 上傳 ZIP 壓縮檔 (內含 JPG/PNG)":
                    with zipfile.ZipFile(uploaded_img, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                else: 
                    try:
                        wb = openpyxl.load_workbook(uploaded_img, data_only=True)
                        sheet_name_to_use = "Products" if "Products" in wb.sheetnames else ("Data" if "Data" in wb.sheetnames else wb.sheetnames[0])
                        sheet = wb[sheet_name_to_use]
                        image_loader = SheetImageLoader(sheet)

                        header_row, thumb_col, dpci_col = None, None, None
                        for r in range(1, 15): 
                            for c in range(1, sheet.max_column + 1):
                                val = str(sheet.cell(row=r, column=c).value).strip().lower()
                                if val == 'thumbnail':
                                    thumb_col = c
                                    header_row = r
                                elif val in ['dpci', 'pid', 'spark pid']:
                                    dpci_col = c
                            if thumb_col and dpci_col:
                                break

                        if thumb_col and dpci_col:
                            thumb_letter = get_column_letter(thumb_col)
                            for r in range(header_row + 1, sheet.max_row + 1):
                                dpci_val = str(sheet.cell(row=r, column=dpci_col).value).strip()
                                img_cell = f"{thumb_letter}{r}"

                                if dpci_val and dpci_val.lower() not in ['none', 'nan', '']:
                                    if image_loader.image_in(img_cell):
                                        try:
                                            img = image_loader.get(img_cell)
                                            safe_name = "".join(x for x in dpci_val if x.isalnum() or x in "-_")
                                            img_path = os.path.join(temp_dir, f"{safe_name}.png")
                                            img.save(img_path, "PNG")
                                        except Exception:
                                            pass
                        else:
                            st.warning("⚠️ 在圖片 Excel 中找不到 'Thumbnail' 標題。")
                    except Exception as e:
                        st.warning(f"⚠️ 解析 Excel 圖片失敗: {e}")
            
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            
            base_props = {
                'label': {'font_name': 'Arial', 'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter'},
                'data': {'font_name': 'Arial', 'font_size': 10, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True},
                'img': {'bg_color': '#FFFFFF'}
            }
            
            fmt = {}
            def create_fmt(name, base, **kwargs):
                p = base_props[base].copy()
                p.update(kwargs)
                fmt[name] = workbook.add_format(p)
                
            create_fmt('img_top', 'img', top=2, left=2, right=2, bottom=1)
            create_fmt('lbl_l', 'label', left=2)
            create_fmt('lbl_in', 'label')
            create_fmt('lbl_lb', 'label', left=2, bottom=2)
            create_fmt('dat_in', 'data')
            create_fmt('dat_r', 'data', right=2)
            create_fmt('dat_b', 'data', bottom=2)
            create_fmt('dat_rb', 'data', right=2, bottom=2)
            
            create_fmt('dat_in_curr', 'data', num_format='$#,##0.00')
            create_fmt('dat_r_curr', 'data', right=2, num_format='$#,##0.00')
            create_fmt('fact_merge', 'label', left=2, right=2, bottom=2, bg_color='#FFC000')

            create_fmt('hdr_title', 'label', font_size=13, align='left')
            create_fmt('hdr_lbl', 'label', align='left')
            create_fmt('hdr_input', 'data', align='left', bg_color='#F4F4F4', border=1)

            def get_val(row_series, possible_cols):
                for col in possible_cols:
                    if col in row_series.index:
                        val = str(row_series[col]).strip()
                        if val.lower() not in ['nan', '', 'nat', 'none']:
                            return val
                return ""
            
            def to_float(val):
                if not val: return ""
                try: return float(val.replace('$', '').replace(',', '').strip())
                except ValueError: return val
                
            def to_int_str_comma(val):
                if not val: return ""
                try: return f"{int(float(str(val).replace(',', '').strip())):,}"
                except ValueError: return str(val)
                
            def w_row(ws, s_row, s_col, r_off, c0, c1, c2, c3, c4, f0, f1, f2, f3, f4):
                ws.write(s_row + r_off, s_col + 0, c0, fmt[f0])
                ws.write(s_row + r_off, s_col + 1, c1, fmt[f1])
                ws.write(s_row + r_off, s_col + 2, c2, fmt[f2]) 
                ws.write(s_row + r_off, s_col + 3, c3, fmt[f3])
                ws.write(s_row + r_off, s_col + 4, c4, fmt[f4])

            df_filtered['RawFactoryName'] = df_filtered.apply(lambda r: get_val(r, ['Factory Name', 'Factory info.', 'Factory']), axis=1)
            df_filtered = df_filtered.sort_values(by='RawFactoryName').reset_index(drop=True)

            def draw_cards_on_sheet(ws, current_df):
                ws.set_landscape()
                ws.set_margins(left=0.3, right=0.3, top=0.4, bottom=0.4)
                ws.fit_to_pages(1, 0)
                
                for i in range(3):
                    base = i * 6
                    ws.set_column(base, base, 13)       
                    ws.set_column(base + 1, base + 1, 22) 
                    ws.set_column(base + 2, base + 2, 2)  
                    ws.set_column(base + 3, base + 3, 13) 
                    ws.set_column(base + 4, base + 4, 22) 
                    ws.set_column(base + 5, base + 5, 4)  

                ws.set_row(0, 30)
                ws.merge_range(0, 0, 0, 1, "202X D240  PROGRAM NAME - ", fmt['hdr_title'])
                ws.merge_range(0, 2, 0, 4, "CATEGORY", fmt['hdr_title'])
                ws.data_validation(0, 2, 0, 4, {'validate': 'list', 'source': ['Decor', 'Kids Activity', 'Party', 'Giveaways', 'TOT']})

                ws.set_row(1, 20)
                ws.merge_range(1, 0, 1, 1, "Business award date:", fmt['hdr_lbl'])
                ws.merge_range(1, 2, 1, 4, "", fmt['hdr_input']) 
                ws.write(1, 6, "Vendor ID#:", fmt['hdr_lbl'])                 
                ws.merge_range(1, 7, 1, 9, "", fmt['hdr_input'])

                ws.set_row(2, 20)
                ws.merge_range(2, 0, 2, 1, "Sourcing:", fmt['hdr_lbl'])
                ws.merge_range(2, 2, 2, 4, "", fmt['hdr_input'])
                ws.data_validation(2, 2, 2, 4, {'validate': 'list', 'source': ['Christy Meyers - Van Der Bosch', 'Ashley Krucker', 'Angela Kennedy']})
                ws.write(2, 6, "PD&D:", fmt['hdr_lbl'])                       
                ws.merge_range(2, 7, 2, 9, "", fmt['hdr_input'])              
                ws.data_validation(2, 7, 2, 9, {'validate': 'list', 'source': ['Adam Hoppus', 'La Dieh Rosenthal', 'Name D']})

                ws.set_row(3, 20)
                ws.merge_range(3, 0, 3, 1, "TSS MR:", fmt['hdr_lbl'])
                ws.merge_range(3, 2, 3, 4, "", fmt['hdr_input'])
                ws.data_validation(3, 2, 3, 4, {'validate': 'list', 'source': ['Asya Yi', 'Jasmine Li', 'Feng Cao']})
                ws.write(3, 6, "Set date:", fmt['hdr_lbl'])                   
                ws.merge_range(3, 7, 3, 9, "", fmt['hdr_input'])

                item_index = 0
                page_breaks = [] 
                
                for index, row in current_df.iterrows():
                    try:
                        block_row = item_index // 3
                        block_col = item_index % 3
                        
                        start_row = 5 + (block_row * 13)
                        start_col = block_col * 6
                        
                        ws.set_row(start_row, 234) 
                        ws.merge_range(start_row, start_col, start_row, start_col + 4, "", fmt['img_top'])
                        
                        for r in range(start_row + 1, start_row + 11):
                            ws.set_row(r, 22)
                        
                        if block_row > 0 and block_row % 2 == 0 and block_col == 0:
                            page_breaks.append(start_row - 1)
                        
                        dpci_val = get_val(row, ['DPCI'])
                        style_val = get_val(row, ['Manufacturer Style # *', 'Manufacturer Style #', 'Style Number'])
                        upc_val = get_val(row, ['Barcode', 'UPC#', 'UPC'])
                        pid_val = get_val(row, ['Spark PID', 'PID']) 
                        desc_val = get_val(row, ['Vendor Product Description *', 'Vendor Product Description', 'Product Description'])
                        
                        # 💡【修改點】防呆消除小數點處理，轉為整數避免 .0 的出現
                        try:
                            if upc_val:
                                upc_val = str(int(float(upc_val)))
                        except Exception:
                            pass
                        
                        fca_val = to_float(get_val(row, ['FCA Factory City Unit Cost', 'FCA', 'FCA $']))
                        retail_val = to_float(get_val(row, ['Suggested Unit Retail', 'RETAIL', 'Retail$']))
                        pack_val = get_val(row, ['Retail Packaging Format (1) *', 'Retail Packaging Format (1)', 'Packaging'])
                        hs_val = get_val(row, ['HTS Code', 'HS NO'])
                        
                        case_q = get_val(row, ['Case Unit Quantity', 'Casepack'])
                        inner_q = get_val(row, ['Inner Pack Unit Quantity', 'Innerpack'])
                        try: case_q = str(int(float(case_q))) if case_q else ""
                        except: pass
                        try: inner_q = str(int(float(inner_q))) if inner_q else ""
                        except: pass
                        pack_str = f"{case_q} / {inner_q}" if (case_q or inner_q) else ""
                        
                        mat_val = get_val(row, ['Primary Raw Material Type', 'Material', 'Main Raw Material *'])
                        qty_val = to_int_str_comma(get_val(row, ['Ent Ttl Rcpt U', 'Total Units', 'QTY']))
                        
                        raw_factory_name = get_val(row, ['Factory Name', 'Factory info.', 'Factory'])
                        raw_factory_id = get_val(row, ['Factory ID', 'Import Vendor ID'])
                        raw_vendor_op = get_val(row, ['Import Vendor Order Point'])
                        op_last_two = raw_vendor_op[-2:] if len(raw_vendor_op) >= 2 else raw_vendor_op
                        
                        fact_parts = [p for p in [raw_factory_name, raw_factory_id, op_last_two] if p]
                        factory_combined = " / ".join(fact_parts)
                        
                        w_row(ws, start_row, start_col, 1, "DPCI:", dpci_val, "", "Style:", style_val,
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        w_row(ws, start_row, start_col, 2, "UPC#:", upc_val, "", "PID:", pid_val,
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        w_row(ws, start_row, start_col, 3, "TCIN:", "", "", "SP:", "☐",
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        
                        ws.data_validation(start_row + 3, start_col + 4, start_row + 3, start_col + 4,
                                           {'validate': 'list', 'source': ['☑', '☐']})
                        ws.write(start_row + 4, start_col, "Description:", fmt['lbl_l'])
                        ws.merge_range(start_row + 4, start_col + 1, start_row + 4, start_col + 4, desc_val, fmt['dat_r'])
                        
                        w_row(ws, start_row, start_col, 5, "FCA $:", fca_val, "", "RETAIL:", retail_val,
                              'lbl_l', 'dat_in_curr', 'dat_in', 'lbl_in', 'dat_r_curr')
                        w_row(ws, start_row, start_col, 6, "Packaging:", pack_val, "", "Red Seal:", "",
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        w_row(ws, start_row, start_col, 7, "HS NO:", hs_val, "", "Casepack:", pack_str,
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        w_row(ws, start_row, start_col, 8, "Material:", mat_val, "", "QTY:", qty_val,
                              'lbl_l', 'dat_in', 'dat_in', 'lbl_in', 'dat_r')
                        
                        ws.write(start_row + 9, start_col, "Remark:", fmt['lbl_l'])
                        ws.merge_range(start_row + 9, start_col + 1, start_row + 9, start_col + 4, "", fmt['dat_r'])
                        ws.merge_range(start_row + 10, start_col, start_row + 10, start_col + 4, f"Factory: {factory_combined}", fmt['fact_merge'])

                        if temp_dir:
                            img_path = None
                            search_names = []
                            for val in [dpci_val, pid_val]:
                                if val:
                                    v_lower = val.lower()
                                    safe_v = "".join(x for x in val if x.isalnum() or x in "-_").lower()
                                    search_names.extend([f"{v_lower}.jpg", f"{v_lower}.png", f"{safe_v}.jpg", f"{safe_v}.png", f"{v_lower}.jpeg", f"{safe_v}.jpeg"])
                            
                            search_names = set(search_names) 
                            
                            for root, dirs, files in os.walk(temp_dir):
                                for file in files:
                                    if file.lower() in search_names:
                                        img_path = os.path.join(root, file)
                                        break
                                if img_path: break
                                
                            if img_path:
                                try:
                                    with Image.open(img_path) as img:
                                        img_w, img_h = img.size
                                    
                                    MAX_BOX_WIDTH = 510
                                    MAX_BOX_HEIGHT = 300
                                    
                                    scale_w = MAX_BOX_WIDTH / img_w
                                    scale_h = MAX_BOX_HEIGHT / img_h
                                    
                                    final_scale = min(scale_w, scale_h)
                                    
                                    final_w = img_w * final_scale
                                    final_h = img_h * final_scale
                                    
                                    x_offset = int((530 - final_w) / 2)
                                    y_offset = int((312 - final_h) / 2)
                                    
                                    ws.insert_image(start_row, start_col, img_path, 
                                                    {'x_scale': final_scale, 'y_scale': final_scale, 
                                                     'x_offset': max(x_offset, 5), 'y_offset': max(y_offset, 5)})
                                except Exception as e:
                                    pass
                        
                        item_index += 1
                    except Exception as e:
                        continue 
                
                if page_breaks:
                    ws.set_h_pagebreaks(page_breaks)

            # ========================================================
            # 🚀 【自動生成頁籤流程】
            # ========================================================
            factory_count = len(df_filtered['RawFactoryName'].unique())
            
            ws_master = workbook.add_worksheet('Master Sheet')
            draw_cards_on_sheet(ws_master, df_filtered)
            
            ws_master.write(1, 15, "Factory#:", fmt['hdr_lbl'])
            ws_master.write(1, 16, factory_count, fmt['hdr_input'])
            ws_master.write(2, 15, "Item#:", fmt['hdr_lbl'])
            ws_master.write(2, 16, len(df_filtered), fmt['hdr_input'])
            
            used_sheet_names = set(['Master Sheet'])
            
            for factory, group_df in df_filtered.groupby('RawFactoryName'):
                clean_name = str(factory).strip()
                if not clean_name: clean_name = "Unknown Factory"
                clean_name = re.sub(r'[\\/*?:\[\]]', '_', clean_name)[:28]
                
                final_name = clean_name
                counter = 1
                while final_name.lower() in [s.lower() for s in used_sheet_names]:
                    final_name = f"{clean_name}_{counter}"
                    counter += 1
                
                used_sheet_names.add(final_name)
                
                ws_fact = workbook.add_worksheet(final_name)
                draw_cards_on_sheet(ws_fact, group_df)

            workbook.close()
            
            st.success(f"排版完成！總表共包含 {len(df_filtered)} 筆商品，並已自動為您拆分成 {factory_count} 個工廠專屬頁籤。")
            st.download_button(
                label="📥 點此下載最新 Program Sheet",
                data=output.getvalue(),
                file_name="Program_Sheet_Auto.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
